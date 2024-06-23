import openpyxl


wb = openpyxl.load_workbook('Заявки.xlsx')
sheet = wb['Sheet1']
MANAGERS = {}


def calculate_salary(total_amount):
    return total_amount/3 - 28000


def add(key, com):
    if key in MANAGERS.keys():
        MANAGERS[key] += com
    else:
        MANAGERS[key] = com


def add_to_dict(key, com):
    if com is None:
        com = 0
    manager_lst = list(key.split(", "))
    if com != 0 and len(manager_lst) > 1:
        com = com / len(manager_lst)
    for name in manager_lst:
        add(name, com)


for i in range(2, sheet.max_row+1):
    manager_name = sheet.cell(row=i, column=5).value
    commission = sheet.cell(row=i, column=10).value
    add_to_dict(manager_name, commission)


new_sheet = wb.create_sheet('Комиссия')
new_sheet["A1"] = "Менеджер"
new_sheet["B1"] = "Комиссия"
new_sheet["C1"] = "Зарплата"

counter = 1
for key, value in MANAGERS.items():
    counter += 1
    salary = calculate_salary(value)
    new_sheet.cell(row=counter, column=1, value=key)
    new_sheet.cell(row=counter, column=2, value=value)
    new_sheet.cell(row=counter, column=3, value=salary)

wb.save('Заявки.xlsx')
